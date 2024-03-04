import os
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re

def replace_word_in_file(file_path, old_word, new_word):
    with open(file_path, 'r') as f:
        content = f.read()

    content = content.replace(old_word, new_word)

    with open(file_path, 'w') as f:
        f.write(content)

def parse_xaml_files(directory):
    ns = {'c': 'clr-namespace:Wdh.Phoenix.CommonUI.Core;assembly=Phoenix.Application.CommonUI.Core'}
    data = []
    exceptions = ['(A)', '(ACT)', '(L)', '(R)', '(US)', 'A/S', 'ABR', 'AC', 'ACT', 'AFBS', 'ASHA', 'ANSI', 'B0', 'BTE', 'BC', 'CIC', 'COSI', 'COIL', 'CR', 'COW', 'CSV', 'CROS', 'DSE', 'DSE,', 'DAI', 'DAI/FM', 'DAI/FM+M', 'DSL', 'DSL5', 'DK', 'DK-2765', 'FM', 'HF', 'HI', 'HI-PRO', 'HI-PRO2', 'HL', 'HCP', 'HTL', 'ICRA', 'IEC', 'IEC-711', 'IG', 'IG,', 'IP', 'ID', 'IIC', 'IIC/CIC', 'ITC', 'ITC/ITE', 'ITEFS', 'ITEHS', 'ISTS', 'ISO', 'ISO/IEC', 'LED', 'LX', 'LF', 'MCL', 'MCR', 'MIC', 'MT', 'MF', 'MF+', 'MF-', 'MPO', 'MINIBTE', 'MINIRITE', 'NAL', 'NAL-NL1', 'NAL-NL2', 'NAL-NL1/NAL-NL2', 'NAL-NL1/NAL-NL2,', 'NOAH', 'NOAH.', 'NOAHLink', 'NOAHlink', 'N0', 'NMFI', 'OSPL90', 'PC', 'PIN', 'PP', 'PRO', 'PSD', 'PX', 'PDF', 'REAR', 'RECD', 'REIR', 'REUG', 'RITE', 'REM', 'SII', 'SPL', 'SPL.', 'SP/UP', 'T-COIL', 'TV', 'UCL', 'US', 'VAC', 'VAC+', 'VAC+,', 'VC', 'XL']

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xaml'):
                try:
                    tree = ET.parse(os.path.join(root, file))
                except ET.ParseError:
                    print(f"Failed to parse {file}")
                    continue
                for element in tree.findall('.//c:PhoenixString', ns):
                    key = element.attrib.get('{http://schemas.microsoft.com/winfx/2006/xaml}Key')
                    original_value = element.text
                    if not original_value:  # If original_value is empty, skip this iteration
                        continue
                    if original_value and isinstance(original_value, str):
                        words = original_value.split()
                        first_upper_found = False
                        changed_words = []
                        for i, word in enumerate(words):
                            # Remove punctuation from the word before comparing
                            word_without_punctuation = word.replace('.', '').replace(',', '')
                            if word_without_punctuation.isupper() and word_without_punctuation not in exceptions:            
                                if not first_upper_found:
                                    words[i] = word.capitalize()
                                    first_upper_found = True, 
                                    changed_words.append((word, words[i]))                                 
                                else:
                                    words[i] = word.lower()
                                replace_word_in_file(os.path.join(root, file), word, words[i])
                        value = ' '.join(words)
                        element.text = value
                        # Determine status
                        if original_value.isupper():
                            status = '1. ALL'
                        elif any(word.isupper() for word in original_value.split()):
                            status = '2. SOME'
                        else:
                            status = '3. NONE'
                    else:
                        status = '3. NONE'
                    data.append([key, original_value, value, status, changed_words])
                ET.register_namespace('', "http://schemas.microsoft.com/winfx/2006/xaml/presentation")
                ET.register_namespace('x', "http://schemas.microsoft.com/winfx/2006/xaml")
                ET.register_namespace('c', "clr-namespace:Wdh.Phoenix.CommonUI.Core;assembly=Phoenix.Application.CommonUI.Core")


    df = pd.DataFrame(data, columns=['Key', 'Original Value', 'Changed Value', 'Status', 'Changed Words'])
    output_file = 'output.xlsx'
    df.to_excel(output_file, index=False)

# Highlight rows based on status
    book = load_workbook(output_file)
    sheet = book.active
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    for row in sheet.iter_rows(min_row=2):
        original_value = row[1].value
        changed_value = row[2].value
        if original_value != changed_value:  # Check if 'Original Value' is not equal to 'Changed Value'
            status = row[3].value
            if status == '1. ALL':
                for cell in row:
                    cell.fill = orange_fill
            elif status == '2. SOME':
                for cell in row:
                    cell.fill = yellow_fill
    book.save(output_file)

parse_xaml_files('c:\\GIT\\Phoenix\\APP\\UserInterface\\Resources\\Genie\\StringLibrary.DesignTime\\')