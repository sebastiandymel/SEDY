import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def read_file(filename):
    with open(filename, 'r') as file:
        lines = file.readlines()
    return lines

def extract_sections(file_lines):
    sections = {}
    current_section = None
    for line in file_lines:
        if line.startswith("###"):
            current_section = line.strip()
            sections[current_section] = []
        elif current_section:
            sections[current_section].append(line.strip())
    return sections

def clean_sheet_name(sheet_name):
    # Replace invalid characters with underscores (you can modify this as needed)
    return sheet_name.replace(":", "_")

def remove_empty_rows(sheet):
    non_empty_rows = [row for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row) if any(cell.value for cell in row)]
    new_sheet = openpyxl.Workbook().active
    for row in non_empty_rows:
        new_sheet.append([cell.value for cell in row])
    return new_sheet

def generate_comparison_report(file1, file2, output_file):
    sections1 = extract_sections(read_file(file1))
    sections2 = extract_sections(read_file(file2))

    wb = openpyxl.Workbook()

    for section in sections1:
        if section in sections2:
            section_name = section.strip("###").strip()
            sheet_name = clean_sheet_name(section_name)

            try:
                ws = wb.create_sheet(title=sheet_name)
            except ValueError:
                # Handle invalid characters in the sheet name
                sheet_name = "Sheet_" + str(wb.sheetnames.index("Sheet") + 1)
                ws = wb.create_sheet(title=sheet_name)

            common_items = set(sections1[section]) & set(sections2[section])

            for r, item in enumerate(common_items, start=1):
                ws.cell(row=r, column=1, value=item)
                ws.cell(row=r, column=2, value=item)

            # Add items that exist in file1 but not in file2
            for r, item in enumerate(sections1[section], start=len(common_items) + 2):
                if item not in common_items:
                    cell = ws.cell(row=r, column=1, value=item)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Add items that exist in file2 but not in file1
            for r, item in enumerate(sections2[section], start=len(common_items) + len(sections1[section]) + 3):
                if item not in common_items:
                    cell = ws.cell(row=r, column=2, value=item)
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Remove empty rows
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        wb.remove(ws)
        wb.create_sheet(sheet)
        wb[sheet] = remove_empty_rows(wb[sheet])

    # Autofit columns
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Add borders to cells
    border = Border(left=Side(border_style="thin"), 
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"))
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    wb.remove(wb['Sheet'])  # Remove the default sheet
    wb.save(output_file)


file1 = "2023-10-25-output.txt"
file2 = "2022-08-29-output.txt"
output_file = "comparison_report.xlsx"
generate_comparison_report(file1, file2, output_file)
